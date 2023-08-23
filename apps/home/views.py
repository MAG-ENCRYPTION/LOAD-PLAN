# -*- encoding: utf-8 -*-
"""
Copyright (c) 2019 - present AppSeed.us
"""
import locale
import os
from datetime import *
from datetime import datetime

import openpyxl
from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.shortcuts import *
from django.template import loader
from django.urls import reverse
from openpyxl.utils import get_column_letter
from openpyxl.styles import *

from .models import *


def ConvertDate(Inputdate):
    date_format = "%Y-%m-%d"
    my_date = datetime.strptime(Inputdate, date_format).date()
    return my_date


def excel_download(request):
    try:
        with open(f'excel/PLAN DE CHARGE {request.user.username}.xlsx', 'rb') as f:
            response = HttpResponse(f.read(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=PLAN DE CHARGE {request.user.username}.xlsx'

        return response
    except Exception:
        return render(request, 'home/page-404.html')


@login_required(login_url="/login/")
def map(request):
    try:
        error = False
        activities = Activity.objects.all()
        UserActs = ActivityUSer.objects.filter(utilisateur_id=request.user.id)
        charges = Charge.objects.filter(id_user_id=request.user.id)
        autoclosed = Cloturation.objects.all()
        UserClosed = CloturationUSer.objects.filter(utilisateur_id=request.user.id)
        projetclos = ProjectFerme.objects.all()
        idSemCLose = []
        ClosedProjetList = []
        for closedProjet in ProjectFerme.objects.all():
            ClosedProjetList.append(closedProjet.nomProjet)
        for cloture in autoclosed:
            idSemCLose.append(cloture.SemaineClose.id.__int__())

        # Récupérer la date actuelle
        today = datetime.now()

        # Calculer la date de début (deux mois avant la date actuelle)
        start_date = today - timedelta(days=60)

        # Calculer la date de fin (deux mois après la date actuelle)
        end_date = today + timedelta(days=60)

        semaines = Semaine.objects.filter(start__range=[start_date, end_date]).order_by('start')
        context = {'acts': activities,
                   'semaines': semaines,
                   'charges': charges,
                   'clotures': autoclosed,
                   'userclotures': UserClosed,
                   'error': error,
                   'UserActs': ActivityUSer.objects.filter(utilisateur_id=request.user.id),
                   'ClosedProjetList': ClosedProjetList,
                   'projetclos': projetclos,
                   'suspendus': WaitingProject.objects.all()
                   }

        if request.method == 'POST':
            EntUSer = EntityUSer.objects.filter(utilisateur=request.user.id).first().entity.id
            idAct = request.POST.get('activity')
            actDeleted = request.POST.get('deleteactivity')
            id_semaine = request.POST.get('semaine')
            charge = request.POST.get('charge')
            SemClose = request.POST.get('semaineclose')
            SemOpen = request.POST.get('semaineouverte')
            projetclos = request.POST.get('closedprojet')
            datecloture = request.POST.get('datecloture')
            waitingprojet = request.POST.get('waitingprojet')
            waitingdate = request.POST.get('waitingdate')
            motif = request.POST.get('motif')
            openedprojet = request.POST.get('openedprojet')
            newdatefinproject = request.POST.get('newdatefinproject')
            nomact = request.POST.get('nomact')
            descriptioinact = request.POST.get('descriptioinact')
            typeact = request.POST.get('typeact')
            startact = request.POST.get('startact')
            endact = request.POST.get('endact')

            # Activité crée par l'utilisateur
            if nomact is not None and descriptioinact is not None and typeact is not None and startact is not None and endact is not None and str(
                    typeact) != 'ras':
                if str(typeact) == 'projet':
                    Activity.objects.create(name=nomact, description=descriptioinact,
                                            type=str(typeact), start=startact, end=endact,
                                            foreverybody=1)
                else:
                    Activity.objects.create(name=nomact, description=descriptioinact,
                                            type=str(typeact), start=startact, end=endact,
                                            foreverybody=0)
                otherAct = Activity.objects.filter(name=nomact).first().id
                ActivityUSer.objects.create(utilisateur_id=request.user.id, activity_id=otherAct)

            # Rouvrir un projet avec une nouvelle date de fermeture.
            if projetclos and datecloture:
                nomprojet = ProjectFerme.objects.filter(id=projetclos).first().nomProjet
                start = ProjectFerme.objects.filter(id=projetclos).first().start
                description = ProjectFerme.objects.filter(id=projetclos).first().description
                Activity.objects.create(name=nomprojet, description=description, type='projet', start=start,
                                        end=datecloture, foreverybody=1)
                ProjectFerme.objects.filter(id=projetclos).delete()

            # Rouvrir un projet suspendu
            if newdatefinproject and openedprojet:
                EstSuspendu = WaitingProject.objects.filter(id=openedprojet).first()
                Activity.objects.create(name=EstSuspendu.nomProjet, type='projet', description=EstSuspendu.description,
                                        start=EstSuspendu.start, end=newdatefinproject, foreverybody=1)
                WaitingProject.objects.filter(id=openedprojet).delete()

            # Suspendre un projet avec motif
            if waitingprojet and waitingdate and motif:
                ASuspendre = Activity.objects.filter(id=waitingprojet).first()
                WaitingProject.objects.create(nomProjet=ASuspendre.name, description=ASuspendre.description,
                                              motif=motif, dateFemeture=waitingdate, start=ASuspendre.start)
                Activity.objects.filter(id=waitingprojet).delete()

            # Supprimer une activité pour un utilisateur
            if actDeleted is not None:
                myactivity = ActivityUSer.objects.filter(id=actDeleted).first().activity.id
                if charges:
                    for charger in charges:
                        if charger.code_activity.id == myactivity:
                            ActivityUSer.objects.filter(id=int(actDeleted)).delete()
                            if ActivityUSer.objects.filter(id=actDeleted) is not None:
                                ActivityUSer.objects.filter(id=actDeleted).delete()
                            if Charge.objects.filter(code_activity_id=charger.code_activity.id) is not None:
                                Charge.objects.filter(code_activity_id=charger.code_activity.id).delete()
                else:
                    ActivityUSer.objects.filter(id=int(actDeleted)).delete()

            # Ajouter une charge pour une activité
            if EntUSer is not None and idAct is not None and id_semaine is not None and charge is not None and charge != '' and is_float(
                    charge) and float(charge) < 7.0 and not Charge.objects.filter(id_semaine_id=id_semaine,
                                                                                  code_activity_id=idAct).exists():
                Charge.objects.create(id_user_id=request.user.id,
                                      id_semaine_id=id_semaine,
                                      code_activity_id=idAct,
                                      charge=float(charge),
                                      modify_date=timezone.now())
                if ActivityUSer.objects.filter(activity_id=idAct, utilisateur_id=request.user.id).first() is None:
                    ActivityUSer.objects.create(utilisateur_id=request.user.id,
                                                activity_id=idAct)
            if SemClose is not None and SemClose != 'default':
                redondance = CloturationUSer.objects.filter(SemaineClose_id=SemClose,
                                                            utilisateur_id=request.user.id).first()
                if redondance is None:
                    instance = CloturationUSer(utilisateur_id=request.user.id, SemaineClose_id=SemClose)
                    instance.save()
                    isclosed = CloturationUSer.objects.filter(utilisateur_id=request.user.id)
                    context = {'acts': activities,
                               'semaines': semaines,
                               'charges': charges,
                               'autoclosed': autoclosed,
                               'error': error,
                               'userclotures': isclosed,
                               'UserActs': ActivityUSer.objects.filter(utilisateur_id=request.user.id),
                               'ClosedProjetList': ClosedProjetList,
                               'projetclos': ProjectFerme.objects.all(),
                               'suspendus': WaitingProject.objects.all()
                               }
            if SemOpen is not None:
                CloturationUSer.objects.filter(utilisateur_id=request.user.id, id=int(SemOpen)).delete()
                isclosed = CloturationUSer.objects.filter(utilisateur_id=request.user.id)
                context = {'acts': activities,
                           'semaines': semaines,
                           'charges': charges,
                           'clotures': isclosed,
                           'autoclosed': autoclosed,
                           'userclotures': UserClosed,
                           'error': error,
                           'UserActs': ActivityUSer.objects.filter(utilisateur_id=request.user.id),
                           'ClosedProjetList': ClosedProjetList,
                           'projetclos': ProjectFerme.objects.all(),
                           'suspendus': WaitingProject.objects.all()
                           }
                return render(request, 'home/map.html', context)
        return render(request, 'home/map.html', context)
    except Exception:
        return render(request, 'home/page-500.html')


def AutoClosedWeek():
    # Obtenir la date actuelle
    now = datetime.now()
    StartYear = datetime(2023, now.month.__int__(), now.day.__int__())
    EndClosed = datetime(2023, 1, 1)

    # Calculer la date un mois avant la date actuelle
    StartYear = StartYear - timedelta(days=60)

    # Récupérer toutes les instances de la classe Semaine qui ont commencé il y a un mois ou moins
    semaines = Semaine.objects.filter(start__lte=StartYear, end__gte=EndClosed)
    # Afficher les instances de la classe Semaine
    for semaine in semaines:
        if Cloturation.objects.filter(SemaineClose_id=semaine.id).first() is None:
            Cloturation.objects.create(SemaineClose_id=semaine.id)


def DateDuJourFormate(DateDuJour):
    DateDuJour = timezone.now().now()
    if DateDuJour.month.__int__() < 10 and DateDuJour.day.__int__() < 10:
        cejour = f'{DateDuJour.year}-0{DateDuJour.month}-0{DateDuJour.day}'
    elif DateDuJour.day.__int__() < 10:
        cejour = f'{DateDuJour.year}-{DateDuJour.month}-0{DateDuJour.day}'
    elif DateDuJour.month.__int__() < 10:
        cejour = f'{DateDuJour.year}-0{DateDuJour.month}-{DateDuJour.day}'
    else:
        cejour = f'{DateDuJour.year}-{DateDuJour.month}-{DateDuJour.day}'
    return cejour


def DateFormate(DateDuJour):
    if DateDuJour.month.__int__() < 10 and DateDuJour.day.__int__() < 10:
        cejour = f'{DateDuJour.year}-0{DateDuJour.month}-0{DateDuJour.day}'
    elif DateDuJour.day.__int__() < 10:
        cejour = f'{DateDuJour.year}-{DateDuJour.month}-0{DateDuJour.day}'
    elif DateDuJour.month.__int__() < 10:
        cejour = f'{DateDuJour.year}-0{DateDuJour.month}-{DateDuJour.day}'
    else:
        cejour = f'{DateDuJour.year}-{DateDuJour.month}-{DateDuJour.day}'
    return cejour


def CloseProject():
    thisDay = DateDuJourFormate(timezone.now().now())
    # On va fermer un projet

    if ActivityUSer.objects.all() is not None:
        for activity in ActivityUSer.objects.all():
            if activity.activity:
                if thisDay == activity.activity.end.__str__().__str__() and activity.activity.type.lower() == 'projet':
                    activity.delete()

        for activity in EntityActivity.objects.all():
            if activity.code_activity.end.__str__().__str__() == thisDay and activity.code_activity.type.lower() == 'projet':
                activity.delete()

        for activity in Activity.objects.all():
            if activity.end.__str__().__str__() == thisDay and activity.type.lower() == 'projet':
                activity.delete()
                ProjectFerme.objects.create(nomProjet=activity.name, start=activity.start, dateFemeture=activity.end,
                                            description=activity.description)


@login_required(login_url="/login/")
def index(request):
    try:
        UpdateChargeDict = {}
        entites = Entite.objects.all()
        SEMAINE = []
        context = {'entity': entites}
        zero = 0
        # Utilisateur par entité
        Personnel = [request.user.id]  # Un membre du personnel est une personne qui appartient à une entité.
        Entuser = EntityUSer.objects.all()
        CloseProject()
        MOI = EntityUSer.objects.filter(utilisateur_id=request.user.id).first()
        """for entuser in Entuser:
                Personnel.append(entuser.utilisateur.id)"""
        context = {'personnel': Personnel, 'entity': entites, 'me': MOI, 'projetclos': ProjectFerme.objects.all()}
        # AutoClosedWeek()
        # Afficher les semaines closes.
        idSemCloses = []
        SemaineCloseQueries = CloturationUSer.objects.filter(utilisateur_id=request.user.id)
        for SemaineCloseQuery in SemaineCloseQueries:
            idSemCloses.append(SemaineCloseQuery.SemaineClose.start)
        convertCloses = []
        for r in idSemCloses:
            convertCloses.append(DateFormate(r))

        AutoSemCloses = []
        AutoSemaineCloseQueries = Cloturation.objects.all()
        for AutoSemaineCloseQuery in AutoSemaineCloseQueries:
            AutoSemCloses.append(AutoSemaineCloseQuery.SemaineClose.start)

        if request.method == 'POST' and request.user.is_authenticated:
            # Récupérer les valeurs du formulaire
            idUser = request.user.id
            charges = Charge.objects.filter(id_user=idUser)
            annee = request.POST.get('annee')
            mois1 = request.POST.get('mois1')
            mois2 = request.POST.get('mois2')
            myentity = request.POST.get('myentity')

            activites = ActivityUSer.objects.filter(utilisateur_id=idUser)

            # Update charge in Table using 'Charge' model in Django Models.
            for charge in charges:
                if request.POST.get(f"{charge.id}") is not None:
                    UpdateChargeDict[f'{charge.id}'] = request.POST.get(f"{charge.id}")
            for cle, valeur in UpdateChargeDict.items():
                if ',' in valeur:
                    valeur = str(valeur)
                    valeur = valeur.replace(',', '.')
                if is_float(valeur):
                    Charge.objects.filter(id=cle).update(charge=valeur)

            # Display entity in all 'select' tag in index.html
            if myentity is not None:
                EntityUSer.objects.create(utilisateur_id=idUser, entity_id=int(myentity))
                return render(request, 'home/index.html', context={'entity': entites, 'me': EntityUSer.objects.filter(
                    utilisateur_id=request.user.id).first()})

            # Organize week with choice of user in the month
            semaines = []
            if annee is not None and mois1 is not None and mois2 is not None:
                WeekCompute(int(annee), int(mois1), int(mois2), semaines)
                for semaine in semaines:
                    SEMAINE.append(semaine.start)

            # Afficher les totaux des charges
            totaux = []
            for semana in SEMAINE:
                x = Semaine.objects.get(start=semana)
                totaux.append(x.totaliser(request.user.id))

            context = {
                'segment': 'index',
                'utilisateur': User.objects.filter(id=request.user.id).first(),
                'personnel': Personnel,
                'userEnt': EntityUSer.objects.all(),
                'entity': entites,
                'charges': charges,
                'semaines': semaines,
                'activites': activites,
                'SEMAINE': SEMAINE,
                'me': EntityUSer.objects.filter(utilisateur_id=request.user.id).first(),
                'zero': zero,
                're': request.POST.keys(),
                'semCloses': idSemCloses,
                'autoCloses': AutoSemCloses,
                'totaux': totaux,
            }
            # Ouvrir le fichier Excel existant
            locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
            workbook = openpyxl.load_workbook('base.xlsx')

            # Sélectionner la feuille active
            worksheet = workbook.active

            # Modifier les données dans le fichier Excel
            if request.user.last_name and request.user.first_name:
                worksheet['A1'] = f'PLAN DE CHARGE {request.user.last_name} {request.user.first_name}'
            else:
                worksheet['A1'] = f'PLAN DE CHARGE {request.user.username.upper()}'
            worksheet['D3'] = f'{DateFormate(SEMAINE[0])}'
            k = 1
            row = 8
            for activite in activites:
                worksheet.cell(row=row, column=2, value=activite.activity.name)
                k = 3
                for sem in semaines:
                    worksheet.cell(row=6, column=k, value=f"""{DateFormate(sem.start)}""")
                    worksheet.cell(row=7, column=k, value=f"""Sem {sem.number}""")
                    charger = Charge.objects.filter(id_user_id=request.user.id, id_semaine_id=sem.id,
                                                    code_activity_id=activite.activity.id).first()
                    if charger:
                        worksheet.cell(row=row, column=k, value=charger.charge)
                    else:
                        worksheet.cell(row=row, column=k, value=0)
                    k += 1
                row += 1
            # Total des charges sur chaque semaine.
            worksheet.cell(row=row, column=2, value=f"TOTAUX")
            c = 3
            for semaine in semaines:
                worksheet.cell(row=row, column=c, value=f"""=SUM(${get_column_letter(c)}$8:${get_column_letter(c)}${row - 1})
                """)
                c += 1

            worksheet.cell(row=7, column=k, value=f"TOTAUX")
            # Total des charges par activités
            d = 8
            for activite in activites:
                worksheet.cell(row=d, column=k,
                               value=f"""=SUM($C${d}:${get_column_letter(k - 1)}${d})""")
                d += 1
            worksheet.cell(row=row, column=k,
                           value=f"""=SUM(${get_column_letter(k)}$8:${get_column_letter(k)}${row-1})""")

            # Formater les cellules en fonction de leur valeur.
            for row in worksheet.iter_rows(min_row=8, max_row=row - 1):
                for cell in row:
                    if cell.value is not None:
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                             bottom=Side(style='thin'))

            # Enregistrer une copie du fichier Excel modifié
            workbook.save(f'excel/PLAN DE CHARGE {request.user.username}.xlsx')

            return render(request, 'home/index.html', context=context)
        return render(request, 'home/index.html', context=context)
    except Exception:
        return render(request, 'home/page-500.html')


def WeekCompute(annee, mois1, mois2, semaines):
    date_debut = datetime(annee, mois1, 1).replace(tzinfo=None)
    date_fin = datetime(annee, mois2, 1).replace(tzinfo=None)
    date_fin += relativedelta(months=1)
    delta = timedelta(days=7)
    while date_debut <= date_fin:
        semaine = Semaine.objects.filter(start__gte=date_debut.replace(tzinfo=None),
                                         end__lte=date_fin.replace(tzinfo=None)).first()
        if semaine:
            semaines.append(semaine)
        date_debut += delta


@login_required(login_url="/login/")
def pages(request):
    context = {}
    try:

        load_template = request.path.split('/')[-1]

        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))
        context['segment'] = load_template

        html_template = loader.get_template('home/' + load_template)
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:

        html_template = loader.get_template('home/page-404.html')
        return HttpResponse(html_template.render(context, request))

    except:
        html_template = loader.get_template('home/page-500.html')
        return HttpResponse(html_template.render(context, request))


def profile(request):
    try:
        tmp = []
        semaines = Semaine.objects.all()
        myactivities = ActivityUSer.objects.filter(utilisateur_id=request.user.id)
        myentity = EntityUSer.objects.filter(utilisateur_id=request.user.id).first()
        for sem in semaines:
            tmp.append(sem.totaliser(userID=request.user.id))
        chargemax = max(tmp)
        chargemin = min(tmp)
        entites = Entite.objects.all()
        entitesUser = EntityUSer.objects.filter()
        context = {'acts': myactivities,
                   'semaines': semaines,
                   'entites': entites,
                   'myentity': myentity,
                   'chargemax': chargemax,
                   'chargemin': chargemin,
                   }
        if request.method == 'POST':
            lastname = request.POST.get('lastname')
            firstname = request.POST.get('firstname')
            MonEntite = request.POST.get('MonEntite')
            username = request.POST.get('username')
            email = request.POST.get('email')
            # Modifier son entité
            if MonEntite is not None:
                EntityUSer.objects.filter(utilisateur_id=request.user.id).update(entity_id=int(MonEntite))

            if username:
                if username is not request.user.username:
                    User.objects.filter(id=request.user.id).update(username=username)
            if firstname:
                if firstname is not request.user.first_name:
                    User.objects.filter(id=request.user.id).update(first_name=firstname)
            if email:
                if email is not request.user.email:
                    User.objects.filter(id=request.user.id).update(email=email)
            if lastname:
                if lastname is not request.user.last_name:
                    User.objects.filter(id=request.user.id).update(last_name=lastname)

            context = {
                'utilisateur': User.objects.filter(id=request.user.id).first(),
                'acts': myactivities,
                'semaines': semaines,
                'entites': entites,
                'myentity': EntityUSer.objects.filter(utilisateur_id=request.user.id).first(),
                'chargemax': chargemax,
                'chargemin': chargemin,
            }
            return render(request, 'home/profile.html', context)
        return render(request, 'home/profile.html', context)
    except Exception:
        return render(request, 'home/page-404.html')


def table(request):
    activities = Activity.objects.all()
    semaines = Semaine.objects.all()
    context = {'acts': activities,
               'semaines': semaines
               }
    return render(request, 'home/tables.html', context)


def is_float(value):
    try:
        float(value)
        return True
    except ValueError:
        return False


